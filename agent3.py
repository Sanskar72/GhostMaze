# Imports
from createGhosts import ghostMoves, initializer
from collections import deque
import numpy as np
import time
import copy
import pandas as pd
from openpyxl import load_workbook

def cleanPath(path):
    """_summary_
        Calculate a cleaner and direct path removing the backtrack and loop iterations
    Args:
        path (list): The path followed by the agent to reach the goal

    Returns:
        _type_: A list with a cleaner and direct path removing the backtrack
    """
    q = deque()
    finalPath = list()
    flag = -1
    for ele in path:
        x,y = ele[0], ele[1]
        if [x,y] not in q:
            q.append([x,y])
        else:
            flag = 1
            while(q):
                popped = q.popleft()
                if popped == [x,y]:
                    q.clear()
                else:
                    finalPath.append([popped[0], popped[1]])
            finalPath.append([x,y])      
    if flag == -1:
        return path
    while q:
        popped = q.popleft()
        finalPath.append([popped[0], popped[1]])
                
    return finalPath


def isValid(a, b, size, grid):
    """_summary_
        Checking for boundary conditions and overall validity
    Args:
        a (int): Integer index declaration
        b (int): Integer index declaration
        size (int): Order of the square matrix used for traversal.
        visited (list): A list to append the visited Locations of the maze
        grid (2D List): The different combinations of 51x51 square matrix/maze generated

    Returns: False based on the condition satisfaction of the degree of the maze & True if the location is visited or blocked
        _type_: Boolean
    """
    # If the cell lies out of bounds
    if (a < 0 or a > size-1) or (b < 0 or b > size-1):
        return False

    # If the cell is already visited
    if grid[a, b] == -1:
        return False
    
    #Otherwise
    return True

def measureDist(x1, y1, grid, ghostGrid, size):
    """_summary_
        Function measure the distance between the agent and the closest ghost to increase survivability
    Args:
        x1 (int): x index position of the agent on the maze
        y1 (int): y index position of the agent on the maze
        grid (2D List): The maze with blocked and unblocked cells along with the ghost spawns
        ghostGrid (2D List): Positions pf the ghosts in the maze
        size (int): Size of the maze

    Returns: Updated Co-ordinates using Distance from agent to the closest ghost.
        _type_: int 
    """
    dir = list()
    dist = list()
    for ghost in ghostGrid:
        # Manhattan Distance
        dir.append([ghost[0]-x1, ghost[1]-y1])
        # Euclidean distance
        dist.append(np.sqrt((ghost[0]-x1)**2 + (ghost[1]-y1)**2))
        
    ind = np.argmin(dist)
    #print("dist", dist)
    #print("dir", dir)
    pos = dir[ind]
    # Compare Manhattan Distance to determine the direction to move to (Up or Down)
    if np.abs(pos[0])<=np.abs(pos[1]):
        if pos[0]>0 and isValid(x1-1, y1, size, grid):
            x1 -= 1
        elif isValid(x1+1, y1, size, grid):
            x1 += 1
    # Compare Manhattan Distance to determine the direction to move to (Left or Right)
    if np.abs(pos[0])>np.abs(pos[1]):
        if pos[1]>0 and isValid(x1, y1-1, size, grid):
            y1 -= 1
        elif isValid(x1, y1+1, size, grid):
            y1 += 1
    
    return x1, y1

def closestGhostDist(x1, y1, ghostGrid):
    """_summary_

    Args:
        x1 (_type_): _description_
        y1 (_type_): _description_
        ghostGrid (_type_): _description_

    Returns:
        _type_: _description_
    """
    dist = list()
    for ghost in ghostGrid:
        # Euclidean Distance
        dist.append(np.sqrt((ghost[0]-x1)**2 + (ghost[1]-y1)**2))
        
    dist = np.min(dist)
    return dist


def planDFS(grid,startX, startY, size):
    """_summary_
        Performing Breadth First Search to reach to the goal block location
    Args:
        grid (2D List): The different combinations of 51x51 square matrix/maze generated
        size (int, optional): Order of the matrix/maze used. Defaults to 51.

    Returns: StatusCode if there is a successful way for the agent to reach its goal/ bottom-right node
        _type_: json
    """
    # Stores indices of the location of the maze cells
    visited = {"00":True}
    childRow = [-1, 0, 1, 0]
    childCol = [0, 1, 0 ,-1]
    
    startQ = deque()
    path = list()
    
    # Mark the starting cell as visited and push it into the goal queue
    startQ.append([startX,startY])
    
    # Iterate while the queue is not empty
    while startQ:
        x1,y1 = startQ.pop()
        path.append([x1,y1])
        if grid[x1,y1] == 10:
            return {"statusCode":200, "path":path}
        # Go to the adjacent blocks on the maze
        for i in range(4):
            childX = x1 + childRow[i]
            childY = y1 + childCol[i]
            if isValid(childX, childY, size, grid) and not visited.get(str(childX)+str(childY), False):
                startQ.append([childX, childY])
                visited[str(childX)+str(childY)] = True
                
                
    return {"statusCode": 400, "path":path}


def executeBFS(grid, size, ghostGrid, prevPosition):
    """_summary_
        Implement the path returned by the plan BFS Function
    Args:
        grid (2D List): _description_The maze with blocked and unblocked cells along with the ghost spawns
        size (int): Shape of the maze
        ghostGrid (list): list of ghost positions on the maze
        prevPosition (list): Stores the data of the cell being occupied by the ghost whether it is blocked or not

    Returns: The path followed by the agent towards the goal
        _type_: list
    """
    x1, y1 = 0, 0
    counter = 0
    safePath = 0
    riskyPath = 0
    finalPath = [[0,0]]
    # Directions to move
    agentMove = [[1,0],[0,1],[0,-1],[-1,0]]
    simulationCount = 7
    while(counter<3000 and [x1,y1] not in ghostGrid):
        if grid[x1,y1] == 10:
            return {"statusCode":200, "path":finalPath, "counter":counter, "safeSimulation": safePath, "riskySimulation":riskyPath}

        #START SIMULATION FOR FUTURE PREDICTION
        finalDict = {"survivalCount": 0, "nextStep": [x1,y1]}
        for ind in range(4):
            tempX, tempY = x1+agentMove[ind][0], y1+agentMove[ind][1]
            if isValid(tempX, tempY, size, grid):
                survivalCount = 0
                for _ in range(simulationCount):
                    tempGrid = copy.deepcopy(grid)
                    tempGhostGrid = copy.deepcopy(ghostGrid)
                    tempPrevPosition = copy.deepcopy(prevPosition)
                    simulBFS = simulateBFS(tempGrid, size, tempGhostGrid, tempPrevPosition, tempX, tempY)
                    #print("Index: ",ind,">>>simulDFS: ",simulBFS)
                    if simulBFS["statusCode"] == 200:
                        survivalCount += 1

                
                if survivalCount>finalDict["survivalCount"]:
                    finalDict["survivalCount"] = survivalCount
                    finalDict["nextStep"] = [tempX, tempY]
                
                print("counter=",counter,"//nextStep=",tempX,",",tempY,"//survivalCount=",survivalCount)
                if finalDict["survivalCount"] >= simulationCount-1:
                    safePath += 1
                    break


        if finalDict["survivalCount"]>1:        
            x1, y1 = finalDict["nextStep"][0], finalDict["nextStep"][1]
        else:#MOVE AWAY FROM CLOSEST GHOSTS
            riskyPath += 1
            x1, y1 = measureDist(x1, y1, grid, ghostGrid, size)
        finalPath.append([x1,y1])
        
        if [x1,y1] in ghostGrid:
            return {"statusCode":400, "path":finalPath, "counter":counter, "safeSimulation": safePath, "riskySimulation":riskyPath}

        #MOVE GHOSTS
        grid, ghostGrid, prevPosition = ghostMoves(grid, ghostGrid, prevPosition)
        counter += 1

        
        print("counter:",counter)
        print("Ghost: ", ghostGrid)
        print("finalDict: ", finalDict)
        print("===================================")

    # print("===================================")
    # print("counter:",counter)
    # print("Ghost: ", ghostGrid)
    # print("Agent: ", x1,y1)
    return {"statusCode":400, "path":finalPath, "counter":counter, "safeSimulation": safePath, "riskySimulation":riskyPath}





def simulateBFS(grid, size, ghostGrid, prevPosition, startX, startY):
    """_summary_
        Simulate all possible paths for the agent to move to find out the most feasible path with the highest survival rate
    Args:
        grid (2D List): _description_The maze with blocked and unblocked cells along with the ghost spawns
        size (int): Shape of the maze
        ghostGrid (list): list of ghost positions on the maze
        prevPosition (list): Stores the data of the cell being occupied by the ghost whether it is blocked or not
        startX (int): X Coordinate
        startY (int): Y Coordinate

    Returns: Final Path from start to goal
        _type_: Json
    """
    # Stores indices of the location of the maze cells
    #visited = np.array([[False]*size]*size)
    #childRow = [-1, 0, 1, 0]
    #childCol = [0, 1, 0 ,-1]
    
    #goalQ.append(grid[goalX, goalY])
    dictBFS = planDFS(grid, startX, startY, size = size)
    statusCode, path = dictBFS.get("statusCode"), dictBFS.get("path")
    path = cleanPath(path)
    if len(path)>0:
        pos = path[0]
        x1, y1 = pos[0], pos[1]
    else:
        x1, y1 = 0, 0
    finalPath = list()
    route = 1
    counter = 0
    replan = 0
    #print(grid)

    # Iterate while the queue is not empty
    while [x1,y1] not in ghostGrid and counter<3000:
        if grid[x1,y1] == 10:
            return {"statusCode":200, "counter":counter, "replanCount":replan}
        
        #BFS PLAN
        if closestGhostDist(x1, y1, ghostGrid)<3:
            x1, y1 = measureDist(x1, y1, grid, ghostGrid, size)
            finalPath.append([x1,y1])
            dictBFS = planDFS(grid, x1, y1, size = size)
            statusCode, path = dictBFS.get("statusCode"), dictBFS.get("path")
            replan += 1
            route = 1
            grid, ghostGrid, prevPosition = ghostMoves(grid, ghostGrid, prevPosition)
            counter += 1
            # print("counter:",counter)
            # print("Ghost: ", ghostGrid)
            # print("Agent: ", x1,y1)
            # print("===============close called===================")
            continue
        

        #AGENT MOVE
        if statusCode == 200:
            path = cleanPath(path)
            # for item in path:
            #     print(item["x"], item["y"])
            pos = path[route]
            x1, y1 = pos[0], pos[1]
            finalPath.append([x1,y1])
            route += 1
            
        elif statusCode == 400: #MOVE AGENT AWAY FROM CLOSEST GHOST
            x1, y1 = measureDist(x1, y1, grid, ghostGrid, size)
            finalPath.append([x1,y1])
        
        if [x1,y1] in ghostGrid:
            return {"statusCode":400, "counter":counter, "replanCount":replan}
            
        #GHOST MOVE
        grid, ghostGrid, prevPosition = ghostMoves(grid, ghostGrid, prevPosition)
        counter += 1
        # if counter%30==0:
        #     print("counter:",counter)
        #     print("Ghost: ", ghostGrid)
        #     print("Agent: ", x1,y1)
        #     print("==================================")
        
    # print("counter:",counter)
    # print("Ghost: ", ghostGrid)
    # print("Agent: ", x1,y1)
    return {"statusCode":400, "counter":counter, "replanCount":replan}


def agent3init(noOfGhosts):
    """_summary_
        Initializer Function for the Agent3
    """
    ghostGrid, grid, prevPosition, size = initializer(noOfGhosts)
    agent3_data = executeBFS(grid, size, ghostGrid, prevPosition)
    data = dict()
    data["StatusCode"] = agent3_data["statusCode"]
    data["Steps"] = agent3_data["counter"]
    data["safeSimulation"] = agent3_data["safeSimulation"]
    data["riskySimulation"] = agent3_data["riskySimulation"]
    return data
    
def dataCollection():
    noOfGhosts = 5
    final_data = list()
    for i in range(1,2):
        tic = time.perf_counter()
        data = agent3init(noOfGhosts)
        toc = time.perf_counter()
        data["time"] = str(toc-tic)
        data["GhostCount"] = noOfGhosts
        final_data.append(data)
        if i%10==0:
            #noOfGhosts += 5
            print("noOfGhosts: ", noOfGhosts)
            
    df1 = pd.DataFrame(final_data)
    book = load_workbook('Agent3.xlsx')
    writer = pd.ExcelWriter('Agent3.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df1.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False, header = False)

    writer.save()
        
dataCollection()