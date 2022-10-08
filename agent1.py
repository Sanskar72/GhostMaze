# Import all the packages and required Functions
from createGhosts import ghostMoves, initializer
from collections import deque
import time
import pandas as pd
from openpyxl import load_workbook

def cleanPath(path):
    """_summary_
        Calculate a cleaner and direct path removing the backtrack and loop iterations
    Args:
        path (_type_): The path followed by the agent to reach the goal

    Returns:
        _type_: A list with a cleaner and direct path removing the backtrack
    """
    q = deque()
    finalPath = list()
    flag = -1
    for ele in path:
        x,y = ele[0], ele[1]
        if [x,y] not in q:
            q.append([x,y])
        else:
            flag = 1
            while(q):
                popped = q.popleft()
                if popped == [x,y]:
                    q.clear()
                else:
                    finalPath.append([popped[0], popped[1]])
            finalPath.append([x,y])      
    if flag == -1:
        return path
    while q:
        popped = q.popleft()
        finalPath.append([popped[0], popped[1]])
                
    return finalPath


def isValid(a, b, size, grid):
    """_summary_
        Checking for boundary conditions and overall validity
    Args:
        a (int): Integer index declaration
        b (int): Integer index declaration
        size (int): Order of the square matrix used for traversal.
        visited (list): A list to append the visited Locations of the maze
        grid (2D List): The different combinations of 51x51 square matrix/maze generated

    Returns: False based on the condition satisfaction of the degree of the maze & True if the location is visited or blocked
        _type_: Boolean
    """
    # If the cell lies out of bounds
    if (a < 0 or a > size-1) or (b < 0 or b > size-1):
        return False

    # If the cell is blocked
    if grid[a, b] == -1:
        return False
    
    #Otherwise
    return True


def planDFS(grid,startX, startY, size):
    """_summary_
        Performing Breadth First Search to reach to the goal block location
    Args:
        grid (2D List): The different combinations of 51x51 square matrix/maze generated
        size (int, optional): Order of the matrix/maze used. Defaults to 51.

    Returns: StatusCode if there is a successful way for the agent to reach its goal/ bottom-right node
        _type_: json
    """
    # Stores indices of the location of the maze cells
    visited = {"00":True}
    childRow = [-1, 0, 1, 0]
    childCol = [0, 1, 0 ,-1]
    
    startQ = deque()
    path = list()
    
    # Mark the starting cell as visited and push it into the goal queue
    startQ.append([startX,startY])
    
    # Iterate while the queue is not empty
    while startQ:
        x1,y1 = startQ.pop()
        path.append([x1,y1])
        if grid[x1,y1] == 10:
            return {"statusCode":200, "path":path}
        # Go to the adjacent blocks on the maze
        for i in range(4):
            childX = x1 + childRow[i]
            childY = y1 + childCol[i]
            if isValid(childX, childY, size, grid) and not visited.get(str(childX)+str(childY), False):
                startQ.append([childX, childY])
                visited[str(childX)+str(childY)] = True
                
                
    return {"statusCode": 400, "path":path}

def executeBFS(grid, size, ghostGrid, prevPosition):
    """_summary_
        Implement the path returned by the plan BFS Function
    Args:
        grid (2D List): _description_The maze with blocked and unblocked cells along with the ghost spawns
        size (int): Shape of the maze
        ghostGrid (list): list of ghost positions on the maze
        prevPosition (list): Stores the data of the cell being occupied by the ghost whether it is blocked or not

    Returns: The path followed by the agent towards the goal
        _type_: list
    """
    # Stores indices of the location of the maze cells
    #visited = np.array([[False]*size]*size)
    #childRow = [-1, 0, 1, 0]
    #childCol = [0, 1, 0 ,-1]
    
    #goalQ.append(grid[goalX, goalY])
    startX, startY = 0, 0
    dictBFS = planDFS(grid, startX, startY, size)
    statusCode, path = dictBFS.get("statusCode"), dictBFS.get("path")
    path = cleanPath(path)
    if len(path)>0:
        pos = path[0]
        x1, y1 = pos[0], pos[1]
    else:
        x1, y1 = 0, 0
    route = 1
    counter = 0

    # Iterate while the queue is not empty
    while [x1,y1] not in ghostGrid and counter<2500:
        if grid[x1,y1] == 10:
            return {"statusCode":200, "path":path, "counter":counter}
        
        #AGENT MOVE
        pos = path[route]
        x1, y1 = pos[0], pos[1]
        route += 1
        
        if [x1,y1] in ghostGrid:
            return {"statusCode":400, "path":path, "counter":counter}
        
        grid, ghostGrid, prevPosition = ghostMoves(grid, ghostGrid, prevPosition)
        counter += 1
        # if counter%100==0:
        #     print("counter:",counter)
        #     print("Ghost: ", ghostGrid)
        #     print("Agent: ", x1,y1)
        #     print("==================================")
    print("counter:",counter)
    print("Ghost: ", ghostGrid)
    print("Agent: ", x1,y1)
    return {"statusCode":400, "path":path, "counter":counter}

def agent1init(noOfGhosts):
    """_summary_
        Initializer for the Agent1
    """
    ghostGrid, grid, prevPosition, size = initializer(noOfGhosts)
    agent1_data = executeBFS(grid, size, ghostGrid, prevPosition)
    data = dict()
    data["StatusCode"] = agent1_data["statusCode"]
    data["Steps"] = agent1_data["counter"]
    
    return data

def dataCollection():
    noOfGhosts =46
    final_data = list()
    for i in range(1,51):
        tic = time.perf_counter()
        data = agent1init(noOfGhosts)
        toc = time.perf_counter()
        data["time"] = str(toc-tic)
        data["GhostCount"] = noOfGhosts
        final_data.append(data)
        if i%10==0:
            noOfGhosts += 1
            print("noOfGhosts: ", noOfGhosts)
            
    df1 = pd.DataFrame(final_data)
    book = load_workbook('Agent1.xlsx')
    writer = pd.ExcelWriter('Agent1.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df1.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False, header = False)

    writer.save()
        
dataCollection()